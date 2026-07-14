import pandas as pd
import openpyxl

print("Loading Self_Performance_Evaluation_Form...")
try:
    file1 = r"C:\Performance_Evaluation\Self_Performance_Evaluation_Form (nabihah) Jan-Jun 2026.xlsx"
    wb1 = openpyxl.load_workbook(file1)
    print("Sheets in Self Performance:", wb1.sheetnames)
    for name in wb1.sheetnames:
        sheet = wb1[name]
        print(f"Sheet '{name}': {sheet.max_row} rows, {sheet.max_column} cols")
except Exception as e:
    print("Error loading file 1:", e)

print("\nLoading Weekly_Report_2026_nabihah_Updated...")
try:
    file2 = r"C:\Performance_Evaluation\Weekly_Report_2026_nabihah_Updated.xlsx"
    wb2 = openpyxl.load_workbook(file2)
    print("Sheets in Weekly Report:", wb2.sheetnames)
    for name in wb2.sheetnames:
        sheet = wb2[name]
        print(f"Sheet '{name}': {sheet.max_row} rows, {sheet.max_column} cols")
except Exception as e:
    print("Error loading file 2:", e)
